"""
    File: Scan_Folder_Extract_Data.py
    Author: Aaron Fortner
    Date: 08/14/2024
    Version: 1.0

    Scan Folder Extract Data

    This Python script scans a specified folder for PDF files, extracts text from the PDFs based on coordinates defined
    in a JSON template, and populates an Excel spreadsheet with the extracted data. The script also moves processed
    files to a designated output folder and logs any files that fail to process.

        Features

        - Folder Scanning: Recursively scans a specified folder and its subfolders for PDF files.
        - Text Extraction: Extracts text from PDF files using coordinates defined in a JSON template.
        - Spreadsheet Population: Populates an Excel spreadsheet with the extracted data.
        - File Management: Moves processed files to a designated output folder.
        - Error Logging: Logs any files that fail to process.

        Requirements

        - Python 3.x
        - `openpyxl` for Excel file manipulation
        - `pymupdf` for PDF text extraction
        - `tkinter` for file and folder selection dialogs

        Usage

        1. Select Folder to Scan: Choose the folder containing the PDF files to be processed.
        2. Select Output Folder: Choose the folder where processed files will be moved.
        3. Select JSON Template: Choose the JSON file containing the coordinates for text extraction.
        4. Run the Script: The script will process the files and populate the spreadsheet.

        Refs

        - https://pymupdf.readthedocs.io/en/latest/index.html
        - https://openpyxl.readthedocs.io/en/stable/index.html
        - https://docs.python.org/3/library/tkinter.html
"""

import os
import re
import json
import shutil  # do not delete, needed for move_file function, commented out for testing
import datetime
import openpyxl
import tkinter as tk
import pymupdf as pmu

from tkinter import filedialog
from openpyxl.styles import Font, colors


# TODO: convert to tkinter dialog?
# TODO: loop until a folder is selected or cancel is clicked
def open_folder_dialog(title, initialdir) -> str:
    """
    Opens a dialog to select a folder.

    Args:
        title(str): The title of the dialog window.
        initialdir(str): The initial directory to open the dialog in.

    Returns:
        selected_folder(str): The path to the selected folder.

    Raises:
        None.
    """

    selected_folder = None
    while not selected_folder:
        selected_folder = filedialog.askdirectory(initialdir=initialdir, title=title, mustexist=True)
        if not selected_folder:
            print("No folder selected. Please select a folder to scan.")
    return selected_folder


# TODO: convert to tkinter dialog?
# TODO: loop until a file is selected or cancel is clicked
def open_file_dialog(title, filetypes, initialdir) -> str:
    """
    Opens a dialog to select a file.

    Args:
        title(str): The title of the dialog window.
        filetypes(list): A list of file types to filter by.
        initialdir(str): The initial directory to open the dialog in.

    Returns:
        file_path(str): The path to the selected file.

    Raises:
        None.
    """

    file_path = None
    while not file_path:
        file_path = filedialog.askopenfilename(initialdir=initialdir,  title=title, filetypes=filetypes)
        if not file_path:
            print("No file selected. Please select a file to proceed.")
    return file_path


# TODO: remove commented out code before production *******************************************************************
def move_file(source_dir, dest_dir, file_name) -> None:
    """
    Moves a file from the 'source_dir' folder to the 'dest_dir' folder.

    Args:
        source_dir(str): The path to the 'To Scan' folder.
        dest_dir(str): The path to the 'Scanned' folder.
        file_name(str): The name of the file to move.

    Returns:
        None

    Raises:
        Exception: If an error occurs moving the file.
    """

    # replace spaces with underscores
    file_name = file_name.replace(" ", "_")

    try:
        print('Moving file...')
        source_path = os.path.join(source_dir, file_name)
        destination_path = os.path.join(dest_dir, file_name)
        # TODO: commented out during testing to keep test files in 'To Scan' folder ***********************************
        # shutil.move(source_path, dest_dir)
        print(f"Moved {file_name} to '{dest_dir}'")
    except Exception as e:
        print(f"An error occurred moving the file: {e}")


# TODO: Alternative idea to create a subfolder for each run of the program and move the files to that subfolder
def create_output_subfolder(base_path: str) -> str:
    """
    Creates a subfolder in the base output path named with the current date and time.

    Args:
        base_path (str): The base output path.

    Returns:
        str: The path to the created subfolder.

    Raises:
        FileNotFoundError: If the subfolder cannot be created.
    """

    current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    subfolder_path = os.path.join(base_path, current_time + '_k-means')
    os.makedirs(subfolder_path, exist_ok=True)
    return subfolder_path


# TODO: add feature to select to include subfolders or not?
def queue_manager(working_directory, output_directory, json_path, sheet) -> None:
    """
    Manages the queue of files in the folder and processes them.

    Args:
        working_directory(str): The path to the folder to scan.
        output_directory(str): The path to the folder to save the scanned files.
        json_path(str): The path to the form template JSON file.
        sheet(Worksheet): The Excel worksheet to populate with the extracted data.

    Returns:
        None

    Raises:
        Exception: If an error occurs processing a file.
    """

    # This will process all files in the folder, including subfolders
    # for root, dirs, files in os.walk(working_directory):

    # Iterate through the files in the folder (not subfolders)
    for filename in os.listdir(working_directory):

        # Create the full path to the file by using os.path.join()
        file_path = os.path.join(working_directory, filename)

        # Check if the path is a file (not a directory)
        if os.path.isfile(file_path):

            # process pdf files
            if file_path.endswith(".pdf"):
                # announce processing file
                print(f"Processing file: {filename}")
                try:
                    # process the file, extract text and populate spreadsheet, then move the file to the scanned folder
                    pdf_processor(file_path, json_path, filename, sheet)
                    move_file(working_directory, output_directory, filename)
                except Exception as e:
                    # log failed file and move to next file
                    print(f"Error processing file: {filename} \n\t{e}")
                    # create file name for log file
                    failed_log_filename = os.path.join(working_directory, "_failed_files.log")
                    with open(failed_log_filename, "a") as log_file:
                        log_file.write(f"{filename}\n")
                    continue
            # skip non-pdf and move to next file
            else:
                print(f"File '{filename}' is not a pdf, skipping.")
                continue
        # skip directories and move to next file
        else:
            # note: REGEX to replace backslashes with forward slashes for windows paths to display correctly, the odd
            #  number of backslashes is due to the escape character for the backslash in python being odd.
            print(f"Path '{re.sub('\\\\', '/', file_path)}' is not a file, skipping.")
            continue

    # save log file of failed files
    if os.path.exists("failed_files.log"):
        with open("failed_files.log", "r") as log_file:
            failed_files = log_file.readlines()
            print("The following files failed to process:")
            for file in failed_files:
                print(file)


# TODO: does this work with multiple 1348s in one pdf? - fixed but doesn't handle PDFs with multiple different forms yet
def pdf_processor(pdf_path, json_path, pdf_name, sheet) -> None:
    """
    Extracts text from a PDF file using the coordinates in a JSON file.

    Args:
        pdf_path(str): The path to the PDF file.
        json_path(str): The path to the JSON file containing the coordinates.
        pdf_name(str): The name of the PDF file.
        sheet(Worksheet): The Excel worksheet to populate with the extracted data.

    Returns:
        None.

    Raises:
        Exception: If an error occurs extracting text from the PDF.
    """

    # Open the PDF file
    doc = pmu.open(pdf_path)

    # If the PDF has multiple pages, process each page separately
    if len(doc) > 1:
        # Iterate through each page in the PDF
        for page_number in range(len(doc)):

            # Load the page
            page = doc.load_page(page_number)

            # Extract text from the PDF using the JSON template
            data = extract_text_from_page(page, json_path)

            # Populate the spreadsheet row with the extracted data for this page
            populate_spreadsheet(data, pdf_name, sheet)

    # If the PDF has only one page, process the page
    else:
        # Extract text from the PDF using the JSON template
        data = extract_text_from_page(doc.load_page(0), json_path)

        # Populate the spreadsheet row with the extracted data for this page
        populate_spreadsheet(data, pdf_name, sheet)

    doc.close()


def extract_text_from_page(pdf_page, json_path) -> list:
    """
    Extracts text from a PDF file using the coordinates in a JSON file.

    Args:
        pdf_page(pmu.Page): The PDF page object.
        json_path(str): The path to the JSON file containing the coordinates.

    Returns:
        extracted_data(list): A list of dictionaries containing the extracted data.

    Raises:
        Exception: If an error occurs extracting text from the PDF.
    """

    # Initialize the list to store the extracted data for this page
    extracted_data = []

    # Load the JSON file containing the coordinates of the fields to extract
    with open(json_path, 'r') as file:
        form_fields = json.load(file)

    # Get the boxes for the first page
    boxes = form_fields["page number: 1"]

    # cycle through the boxes in the json file and extract the text from the pdf for each box
    for box in boxes:
        name = box['name']
        coords = box['coords']
        # Convert coordinates to a PyMuPDF Rect object
        pdf_rect = pmu.Rect(coords[0], coords[1], coords[2], coords[3])

        text = pdf_page.get_textbox(pdf_rect).strip()
        # Remove newline characters and excessive whitespace
        text = ' '.join(text.split())
        extracted_data.append({'name': name, 'text': text})

    # # Print the extracted data in the specified format
    # for item in extracted_data:
    #     print(f"{item['name']} = {item['text']}")

    return extracted_data


# TODO: uncomment before production ************************************************************************************
def populate_spreadsheet(fields, pdf_name, sheet) -> None:
    """
    Populates a spreadsheet with the extracted data from a PDF file.

    Args:
        fields(list): A list of dictionaries containing the extracted data.
        pdf_name(str): The name of the PDF file.
        sheet(Worksheet): The Excel worksheet to populate with the extracted data.

    Returns:
        None

    Raises:
        Exception: If an error occurs populating the spreadsheet.
    """

    # Define the hyperlink font style
    hyper_link_font = Font(color=colors.BLUE, underline='single')

    # Create full file path for hyperlink to the original file
    pdf_name = pdf_name.replace(" ", "_")
    # TODO: uncomment before production *******************************************************************************
    # folder = "./Scanned"
    folder = "./To Scan"
    pdf_link_name = os.path.join(folder, pdf_name)

    # Check if the sheet already has headers
    if sheet.dimensions == "A1:A1":
        # Add headers
        headers = ["filename"] + [item['name'] for item in fields]
        sheet.append(headers)

    # Add extracted field data to the sheet
    row = ['=HYPERLINK("{}","{}")'.format(pdf_link_name, pdf_name),
           # Document name with hyperlink
           fields[0]['text'],  # Document number
           fields[1]['text'],  # Nomenclature
           fields[2]['text'],  # NSN
           fields[3]['text'],  # UI
           fields[4]['text'],  # QTY
           re.sub(' ', '.', fields[5]['text']),  # Unit Price, replace space with decimal
           re.sub(' ', '.', fields[6]['text']),  # Total Price, replace space with decimal
           fields[7]['text'],  # Disposal Authorization Code
           fields[8]['text'],  # DEMIL Code
           fields[9]['text'],  # Supply Condition Code
           fields[10]['text'],  # Shipped From
           fields[11]['text'],  # Shipped To
           fields[12]['text'],  # POC name
           re.sub('Phone ', '', fields[13]['text']),  # POC Phone, strip 'Phone ' from the beginning
           re.sub('Email ', '', fields[14]['text']),  # POC Email, strip 'Email ' from the beginning
           ]

    # Append the row to the sheet
    sheet.append(row)
    # Get the current row count to allow styling of the hyperlink
    count = int(sheet.dimensions[4:])
    # Apply hyperlink style to the pdf link
    sheet.cell(row=count, column=1).font = hyper_link_font

    print(f"Data from {pdf_name} added to the spreadsheet")


def main() -> None:
    """
    Main function to scan a folder and extract data from PDF files.

    Args:
        N/A

    Returns:
        None

    Raises:
        None
    """

    # Select the folder to scan
    to_scan_folder = open_folder_dialog("Select Folder to Scan", "./To Scan")

    # Open the existing workbook or create a new one if it doesn't exist
    if os.path.exists("./Scanned/scanned_data.xlsx"):
        # Load the existing workbook
        workbook = openpyxl.load_workbook("./Scanned/scanned_data.xlsx")
        sheet = workbook.active
    else:
        # Create a new workbook and select the active worksheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

    # TODO: make output folder READ ONLY at the end of process to prevent accidental deletion? Can do this with group
    #  membership and permissions but can we rely on what OS and permissions user has? Same issue with making the folder
    #  hidden and read only for the form template folder as well.
    # Select the folder to save the scanned files to
    scanned_folder = open_folder_dialog("Select Folder to Save Scanned Files to", "./Scanned")

    # Select the form template JSON
    json_path = open_file_dialog("Select Form Template", [("JSON Files", "*.json")],
                                 "./Form Templates")

    # Process the files in the folder through the queue manager
    queue_manager(to_scan_folder, scanned_folder, json_path, sheet)

    # Save the workbook
    workbook.save("./Scanned/scanned_data.xlsx")
    print("Data added to the spreadsheet")
    workbook.close()


if __name__ == '__main__':

    main()
